import load_data
import post_process
import plots
import numpy as np
import os
import openpyxl
from load_data import load_constants

# Load input data for the reference case (no scaling, no BESS, no EVs)
data = load_data.load(reference_case=True)

# Ensure no EVs in reference case
data['ev_demand'] = np.zeros(data['n_steps'])

# Verify required data fields
required_fields = ['pv_power', 'consumer_demand', 'grid_buy_price', 'grid_sell_price']
for field in required_fields:
    if field not in data:
        raise KeyError(f"Required field '{field}' missing from data")

pv_power = data['pv_power']
consumer_demand = data['consumer_demand']

# Initialize result arrays (no BESS, no EVs)
P_PV_consumer_vals = np.zeros(data['n_steps'])
P_PV_grid_vals = np.zeros(data['n_steps'])
P_grid_consumer_vals = np.zeros(data['n_steps'])
SOC_vals = np.zeros(data['n_steps'] + 1)  # remains zero

# Additional zero arrays for consistency with other cases
P_PV_EV_vals = np.zeros(data['n_steps'])
P_BESS_EV_vals = np.zeros(data['n_steps'])
P_grid_EV_vals = np.zeros(data['n_steps'])
P_BESS_consumer_vals = np.zeros(data['n_steps'])
P_PV_BESS_vals = np.zeros(data['n_steps'])
P_grid_BESS_vals = np.zeros(data['n_steps'])
P_BESS_grid_vals = np.zeros(data['n_steps'])
P_BESS_charge_vals = np.zeros(data['n_steps'])
P_BESS_discharge_vals = np.zeros(data['n_steps'])

total_net_per_step = np.zeros(data['n_steps'])

# Simulate energy cascade: PV to consumer, excess PV to grid, deficit from grid
for t in range(data['n_steps']):
    pv_t = pv_power[t]
    demand_t = consumer_demand[t]
    buy_t = data['grid_buy_price'][t]
    sell_t = data['grid_sell_price'][t]

    # PV used for consumer demand first
    used_pv_consumer = min(pv_t, demand_t)
    surplus_pv = max(pv_t - used_pv_consumer, 0)
    remaining_consumer_demand = max(demand_t - used_pv_consumer, 0)

    # Assign values
    P_PV_consumer_vals[t] = used_pv_consumer
    P_PV_grid_vals[t] = surplus_pv
    P_grid_consumer_vals[t] = remaining_consumer_demand

    # Compute revenue for each time step
    pv_grid = surplus_pv * sell_t * data['delta_t']
    grid_cost = - remaining_consumer_demand * buy_t * data['delta_t']
    total_net_per_step[t] = pv_grid + grid_cost

# Compile results dictionary for plotting and post-processing
results = {
    'P_PV_gen': pv_power,  # Add this to enable correct flow prioritization in post_process
    'P_PV_consumer_vals': P_PV_consumer_vals,
    'P_PV_grid_vals': P_PV_grid_vals,
    'P_grid_consumer_vals': P_grid_consumer_vals,
    'SOC_vals': SOC_vals,
    'P_grid_sold': P_PV_grid_vals,
    'P_grid_bought': P_grid_consumer_vals,
    'P_PV_EV_vals': P_PV_EV_vals,
    'P_BESS_EV_vals': P_BESS_EV_vals,
    'P_grid_EV_vals': P_grid_EV_vals,
    'P_BESS_consumer_vals': P_BESS_consumer_vals,
    'P_PV_BESS_vals': P_PV_BESS_vals,
    'P_grid_BESS_vals': P_grid_BESS_vals,
    'P_BESS_grid_vals': P_BESS_grid_vals,
    'P_BESS_charge_vals': P_BESS_charge_vals,
    'P_BESS_discharge_vals': P_BESS_discharge_vals
}

results['P_BESS_discharge'] = np.zeros(data['n_steps'])
results['P_BESS_charge'] = np.zeros(data['n_steps'])

# Compute revenues and print summary results
revenues = post_process.compute_revenues(results, data)
post_process.print_results(revenues, results, data)


# Calculate total cost if no PV plant
total_revenue = revenues['total_revenue']  # € over simulation period
total_no_pv_energy_cost = np.sum(consumer_demand * data['grid_buy_price'] * data['delta_t'])
net_energy_cost_with_pv = -total_revenue  # since total_revenue = rev - cost, net cost = cost - rev = -total_revenue
print(f"Net energy cost with PV: {net_energy_cost_with_pv} €")
print(f"Net energy cost without PV: {total_no_pv_energy_cost} €")


# Compute key metrics for LCOE/Financial Model (in kWh, using delta_t in hours; simplified, no BESS)
delta_t = data['delta_t']
total_pv_energy = np.sum(results['P_PV_gen']) * delta_t  # Total PV produced (kWh) over simulation period
total_grid_sold = np.sum(results['P_grid_sold']) * delta_t  # Grid export (kWh)
total_grid_bought = np.sum(results['P_grid_bought']) * delta_t  # Grid import (kWh)
self_sufficiency = revenues['self_sufficiency']  # % (renewable coverage of consumer demand)
ev_renewable_share = revenues['ev_renewable_share']  # % (renewable coverage of EV demand)
total_bess_discharge = 0  # No BESS in reference case
bess_capacity = 0  # No BESS in reference case
pv_old = float(load_constants()['PV_OLD'])
pv_new = float(load_constants()['PV_NEW'])
pv_scaling_factor = (pv_new + pv_old) / pv_old if pv_old > 0 else 1  # Likely 1 for reference, but computed for consistency
bess_to_grid_revenue = revenues['total_bess_to_grid_rev']  # 0 in reference case
bess_to_ev_revenue = revenues['total_bess_to_ev_rev']  # 0 in reference case
pv_to_grid_revenue = revenues['total_pv_to_grid_rev']  # Revenue from PV to Grid
pv_to_ev_revenue = revenues['total_pv_to_ev_rev']  # 0 in reference case
grid_import_cost = revenues['total_grid_buy_cost']  # Cost from Grid Import


# Other relevant parameters (no BESS)
pv_old = float(load_constants()['PV_OLD'])
pv_new = float(load_constants()['PV_NEW'])
pv_scaling_factor = (pv_new + pv_old) / pv_old if pv_old > 0 else 1
simulation_days = 7  # Assuming 7-day simulation; adjust if different for extrapolation in Excel

# Organize data as a dictionary (simplified, no BESS keys)
export_data = {
'REFERENCE': 'Original PV, no BESS, no EVs',
'Total PV Energy Produced (kWh)': total_pv_energy,
'Total BESS Energy Discharged (kWh)': total_bess_discharge,
'Total Grid Sold (kWh)': total_grid_sold,
'Total Grid Bought (kWh)': total_grid_bought,
'Self-Sufficiency Ratio (%)': self_sufficiency,
'EV Renewable Share (%)': ev_renewable_share,
'Total Revenue (€)': total_revenue,
'BESS Capacity (kWh)': bess_capacity,
'PV Scaling Factor': pv_scaling_factor,
'Simulation Period (days)': simulation_days,
'Revenue BESS to Grid (€)': bess_to_grid_revenue,
'Revenue BESS to EV (€)': bess_to_ev_revenue,
'Revenue PV to Grid (€)': pv_to_grid_revenue,
'Revenue PV to EV (€)': pv_to_ev_revenue,
'Cost Grid Import (with small PV) (€)': grid_import_cost,
'Net Energy Cost with PV (€)': net_energy_cost_with_pv,
'Net Energy Cost without PV (€)': total_no_pv_energy_cost
}
# Path to your existing Excel file
excel_path = r'C:\Users\dell\V1_First_Model\Input Data Files\Financial_Model.xlsx'

# Load existing workbook
wb = openpyxl.load_workbook(excel_path)

# Create or select sheet for inputs (won't overwrite other sheets)
sheet_name = 'Output PyPSA'
if sheet_name not in wb.sheetnames:
    ws = wb.create_sheet(sheet_name)
else:
    ws = wb[sheet_name]
    # Clear only columns 1 and 2 (A and B) for all rows with data
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.value = None

# Write data to sheet (Column A: keys, Column B: values; starting at row 1)
row = 1
for key, value in export_data.items():
    ws.cell(row=row, column=1, value=key)
    ws.cell(row=row, column=2, value=value)
    row += 1

# Save updated workbook
wb.save(excel_path)
print(f"Data exported to {excel_path} in sheet '{sheet_name}'")

# Generate plots with _Reference suffix
days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
data['day_labels'] = [days[(data['start_weekday'] + d) % 7] for d in range(7)]
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Output Files')
os.makedirs(output_dir, exist_ok=True)
data['plot_suffix'] = '_Reference'
plots.plot_energy_flows(results, data, revenues, save_dir=output_dir)
plots.plot_financials(revenues, data, save_dir=output_dir)