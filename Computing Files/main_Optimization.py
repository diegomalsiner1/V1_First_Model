# Main optimization script for energy system simulation.
# Loads input data, runs MPC optimization, collects results, and generates plots.

import time
from Controller.mpc import MPC
from Controller.arbitrage_controller import ArbitrageController
from Controller.load_data import load_constants
import Controller.load_data as load_data
import numpy as np
import PostPlot.post_process as post_process
import PostPlot.plots as plots
import sys
import os
import openpyxl
import pandas as pd  # Import pandas for handling datetime

# Start timing
start_time = time.time()
print(f"Starting optimization at {time.strftime('%H:%M:%S')}")

# Print Python executable path for debugging environment issues
print(sys.executable)

# Load all input data (from CSV or API, depending on load_data settings)
data = load_data.load(reference_case=False, price_source="HPFC", base_forecast=100.0, peak_forecast=120.0)
data_load_time = time.time()
print(f"Data loading completed in {data_load_time - start_time:.2f} seconds")
# Debug flag for verbose output
DEBUG = True

# Validate input data: ensure all required keys are present
required_data_keys = [
    'pv_power', 'consumer_demand', 'ev_demand', 'grid_buy_price',
    'grid_sell_price', 'pi_ev', 'lcoe_pv', 'n_steps', 'delta_t',
    'eta_charge'  # Add BESS efficiency for post-processing
]
missing_keys = [key for key in required_data_keys if key not in data]
if missing_keys:
    raise KeyError(f"Missing required keys in data: {missing_keys}")

if DEBUG:
    print("Data validation passed")
    print(f"Number of timesteps: {data['n_steps']}")
    print(f"EV price loaded: {data.get('pi_ev', 'MISSING')}")
    print(f"BESS charge efficiency: {data.get('eta_charge', 'MISSING')}")
    print(f"BESS discharge efficiency: {data.get('eta_discharge', 'MISSING')}")

# Assert consistency of data length for all main time series
assert len(data['pv_power']) == data['n_steps']
assert len(data['consumer_demand']) == data['n_steps']
assert len(data['ev_demand']) == data['n_steps']
assert len(data['grid_buy_price']) == data['n_steps']
assert len(data['grid_sell_price']) == data['n_steps']

# Controller selection
horizon = 672  # Forecast horizon: 7 days (15-min steps)
controller_type = str(load_data.load_constants().get('CONTROLLER_TYPE', 'ARBITRAGE')).upper()  #MPC or ARBITRAGE
if controller_type == 'ARBITRAGE':
    _const = load_data.load_constants()
    # Optional tunables for arbitrage controller (fall back to defaults if missing)
    window_hours = float(_const.get('ARBI_WINDOW_HOURS', 24.0))
    gap_min = float(_const.get('ARBI_GAP_MIN', 0.0))
    alpha_soc_charge = float(_const.get('ARBI_ALPHA_SOC_CHARGE', 0.0))
    beta_soc_discharge = float(_const.get('ARBI_BETA_SOC_DISCHARGE', 0.0))
    gamma_pv_inflow = float(_const.get('ARBI_GAMMA_PV_INFLOW', 0.0))
    hold_steps = int(float(_const.get('ARBI_HOLD_STEPS', 0)))
    term_soc_raw = _const.get('ARBI_TERMINAL_SOC_RATIO', None)
    try:
        terminal_soc_ratio = None if term_soc_raw is None else float(term_soc_raw)
    except Exception:
        terminal_soc_ratio = None

    mpc_controller = ArbitrageController(
        delta_t=data['delta_t'],
        bess_capacity=data['bess_capacity'],
        bess_power_limit=data['bess_power_limit'],
        eta_charge=data['eta_charge'],
        eta_discharge=data['eta_discharge'],
        window_hours=window_hours,
        gap_min=gap_min,
        alpha_soc_charge=alpha_soc_charge,
        beta_soc_discharge=beta_soc_discharge,
        gamma_pv_inflow=gamma_pv_inflow,
        hold_steps=hold_steps,
        terminal_soc_ratio=terminal_soc_ratio
    )
else:
    mpc_controller = MPC(delta_t=data['delta_t'])

# Initialize arrays for storing results
n_steps = data['n_steps']
soc_actual = np.zeros(n_steps + 1)
soc_actual[0] = data['soc_initial']
P_PV_consumer_vals = np.zeros(n_steps)
P_PV_ev_vals = np.zeros(n_steps)
P_PV_grid_vals = np.zeros(n_steps)
P_BESS_discharge_vals = np.zeros(n_steps)
P_BESS_charge_vals = np.zeros(n_steps)
P_grid_consumer_vals = np.zeros(n_steps)
P_grid_ev_vals = np.zeros(n_steps)
P_Grid_to_BESS_vals = np.zeros(n_steps)
P_grid_import_vals = np.zeros(n_steps)
P_grid_export_vals = np.zeros(n_steps)
P_PV_gen = np.zeros(n_steps)  # PV generation profile

# Helper function: pad forecast arrays to match horizon length
def pad_to_horizon(arr, horizon, pad_value=None):
    if len(arr) < horizon:
        if len(arr) > 0:
            return np.pad(arr, (0, horizon - len(arr)), mode='edge')
        else:
            return np.zeros(horizon)
    return arr[:horizon]

# Main MPC loop: run optimization for each timestep
for t in range(n_steps):
    pv_forecast = pad_to_horizon(data['pv_power'][t:t + horizon], horizon)
    demand_forecast = pad_to_horizon(data['consumer_demand'][t:t + horizon], horizon)
    ev_forecast = pad_to_horizon(data['ev_demand'][t:t + horizon], horizon)
    buy_forecast = pad_to_horizon(data['grid_buy_price'][t:t + horizon], horizon)
    sell_forecast = pad_to_horizon(data['grid_sell_price'][t:t + horizon], horizon)

    # Calculate the correct start_dt for this forecast window
    current_start_dt = data['start_dt'] + pd.Timedelta(minutes=15*t)

    control = mpc_controller.predict(
        soc_actual[t], pv_forecast, demand_forecast, ev_forecast,
        buy_forecast, sell_forecast, data['lcoe_pv'], data['pi_ev'], data['pi_consumer'], horizon, current_start_dt
    )

    # Store results for each timestep
    P_PV_consumer_vals[t] = control['pv_bess_to_consumer']
    P_PV_ev_vals[t] = control['pv_bess_to_ev']
    P_PV_grid_vals[t] = control['pv_bess_to_grid']
    P_BESS_discharge_vals[t] = control['P_BESS_discharge']
    P_BESS_charge_vals[t] = control['P_BESS_charge']
    P_grid_consumer_vals[t] = control['grid_to_consumer']
    P_grid_ev_vals[t] = control['grid_to_ev']
    P_Grid_to_BESS_vals[t] = control['P_grid_to_bess']
    P_grid_import_vals[t] = control['P_grid_import']
    P_grid_export_vals[t] = control['P_grid_export']
    soc_actual[t + 1] = control['SOC_next']
    P_PV_gen[t] = control['P_PV_gen']
    

# MPC loop completed
mpc_complete_time = time.time()
print(f"MPC optimization completed in {mpc_complete_time - data_load_time:.2f} seconds")
print(f"Average time per timestep: {(mpc_complete_time - data_load_time)/n_steps:.3f} seconds")

# --- DEBUG: Print first 48 values of key arrays for inspection ---
print("\n--- DEBUG: First 48 values of key dispatch/result arrays ---")
print("P_PV_consumer_vals:", P_PV_consumer_vals[:48])
print("P_BESS_charge_vals:", P_BESS_charge_vals[:48])
print("P_BESS_discharge_vals:", P_BESS_discharge_vals[:48])
print("P_grid_consumer_vals:", P_grid_consumer_vals[:48])
print("P_grid_import_vals:", P_grid_import_vals[:48])
print("SOC_vals:", soc_actual[:49])  # 49 to show initial + 48 steps
print("P_PV_gen:", P_PV_gen[:48])
print("P_PV_grid_vals:", P_PV_grid_vals[:48])
print("P_Grid_to_BESS_vals:", P_Grid_to_BESS_vals[:48])
print("----------------------------------------------------------\n")

# Compile results for post-processing and plotting
results = {
    'P_PV_consumer_vals': P_PV_consumer_vals,
    'P_PV_ev_vals': P_PV_ev_vals,
    'P_PV_grid_vals': P_PV_grid_vals,
    'P_BESS_discharge': P_BESS_discharge_vals,
    'P_BESS_charge': P_BESS_charge_vals,
    'P_grid_consumer_vals': P_grid_consumer_vals,
    'P_grid_ev_vals': P_grid_ev_vals,
    'P_Grid_to_BESS': P_Grid_to_BESS_vals,
    'P_grid_import_vals': P_grid_import_vals,
    'P_grid_export_vals': P_grid_export_vals,
    'SOC_vals': soc_actual,
    'P_PV_gen': P_PV_gen,
    # Note: P_grid_sold and P_BESS_grid_vals will be calculated in post_process.compute_revenues()
    # These preliminary values will be overridden by the post-processing function
    'P_grid_sold': P_PV_grid_vals,  # Preliminary - will be corrected to include BESS exports in post-processing
    'P_grid_bought': P_grid_consumer_vals + P_grid_ev_vals + P_Grid_to_BESS_vals
}

# Compute revenues and print summary results
revenues = post_process.compute_revenues(results, data)
post_process.print_results(revenues, results, data)
post_process_time = time.time()
print(f"Post-processing completed in {post_process_time - mpc_complete_time:.2f} seconds")

#INSERT FROM HERE
# Compute key metrics for LCOE/Financial Model (in kWh, using delta_t in hours)
delta_t = data['delta_t']
total_pv_energy = np.sum(results['P_PV_gen']) * delta_t  # Total PV produced (kWh) over simulation period
total_bess_discharge = np.sum(results['P_BESS_discharge']) * delta_t  # Total BESS provided energy (kWh)
total_pv_to_consumer = np.sum(results['P_PV_consumer_vals']) * delta_t  # PV direct to consumer (kWh)
total_bess_to_consumer = np.sum(results['P_BESS_discharge']) * delta_t  # BESS discharge to consumer (kWh)
total_grid_sold = np.sum(results['P_grid_sold']) * delta_t  # Grid export (kWh)
total_grid_bought = np.sum(results['P_grid_bought']) * delta_t  # Grid import (kWh)
self_sufficiency = revenues['self_sufficiency']  # % (renewable coverage of consumer demand)
ev_renewable_share = revenues['ev_renewable_share']  # % (renewable coverage of EV demand)
total_revenue = revenues['total_revenue']  # € over simulation period

# Other relevant parameters
bess_capacity = data['bess_capacity']  # kWh
pv_old = float(data.get('pv_old', 0))  # Get from data if available, otherwise 0
pv_new = float(data.get('pv_new', 0))  # Get from data if available, otherwise 0
pv_scaling_factor = (pv_new + pv_old) / pv_old if pv_old > 0 else 1
simulation_days = 7  # Assuming 7-day simulation; adjust if different for extrapolation in Excel
bess_to_grid_revenue = revenues['total_bess_to_grid_rev']  # Revenue from BESS to grid
bess_to_ev_revenue = revenues['total_bess_to_ev_rev']      # Revenue from BESS to EV
pv_to_grid_revenue = revenues['total_pv_to_grid_rev']      # Revenue from PV to grid
pv_to_ev_revenue = revenues['total_pv_to_ev_rev']          # Revenue from PV to EV
grid_import_cost = revenues['total_grid_buy_cost']         # Cost by grid import

# Organize data as a dictionary (add more keys if needed, e.g., for CAPEX sensitivity)
export_data = {
    'OPTIMIZED': 'BIG PV, BESS, EVs',
    'Total PV Energy Produced (kWh)': total_pv_energy,
    'Total BESS Energy Discharged (kWh)': total_bess_discharge,
    'Total Grid Sold (kWh)': total_grid_sold,
    'Total Grid Bought (kWh)': total_grid_bought,
    'Self-Sufficiency Ratio (%)': self_sufficiency,
    'EV Renewable Share (%)': ev_renewable_share,
    'Total Revenue (€)': total_revenue,
    'BESS Capacity (kWh)': bess_capacity,
    'PV Scaling Factor': pv_scaling_factor,
    'Simulation Period (days)': simulation_days,
    'Revenue BESS to Grid (€)': bess_to_grid_revenue,
    'Revenue BESS to EV (€)': bess_to_ev_revenue,
    'Revenue PV to Grid (€)': pv_to_grid_revenue,
    'Revenue PV to EV (€)': pv_to_ev_revenue,
    'Cost Grid Import (€)': grid_import_cost,
    'PV to Consumer (kWh)': total_pv_to_consumer,
    'BESS to Consumer (kWh)': total_bess_to_consumer

}

# Path to your existing Excel file (updated for current system)
excel_path = r'C:\Users\diego\V1_First_Model-1\Input Data Files\Financial_Model.xlsx'

# Load existing workbook
wb = openpyxl.load_workbook(excel_path)

# Create or select sheet for inputs (won't overwrite other sheets)
sheet_name = 'Output PyPSA'
if sheet_name not in wb.sheetnames:
    ws = wb.create_sheet(sheet_name)
else:
    ws = wb[sheet_name]
    # Clear only columns 3 and 4 (C and D) for all rows with data
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=4):
        for cell in row:
            cell.value = None

# Write data to sheet (Column C: keys, Column D: values; starting at row 1)
row = 1
for key, value in export_data.items():
    ws.cell(row=row, column=3, value=key)
    ws.cell(row=row, column=4, value=value)
    row += 1

# Save updated workbook
wb.save(excel_path)
print(f"Data exported to {excel_path} in sheet '{sheet_name}'")
#END EXCEL SAVE AND EXPORT


# Prepare output directory for plots
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Output Files')
os.makedirs(output_dir, exist_ok=True)
data['plot_suffix'] = ''  # No suffix for main optimization

# Add grid import/export data to data dictionary for arbitrage visualization
data['grid_import_vals'] = P_grid_import_vals
data['grid_export_vals'] = P_grid_export_vals

# Prepare day labels for plotting (for x-axis ticks)
days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
data['day_labels'] = [days[(data['start_weekday'] + d) % 7] for d in range(7)]

# Generate plots for energy flows and financials
plots.plot_energy_flows(results, data, revenues, save_dir=output_dir)
plots.plot_financials(revenues, data, save_dir=output_dir)
plotting_time = time.time()
print(f"Plotting completed in {plotting_time - post_process_time:.2f} seconds")

# End timing and print results
end_time = time.time()
runtime = end_time - start_time
print(f"\n{'='*50}")
print(f"OPTIMIZATION COMPLETED SUCCESSFULLY!")
print(f"Runtime: {runtime:.2f} seconds ({runtime/60:.2f} minutes)")
print(f"Finished at: {time.strftime('%H:%M:%S')}")
print(f"{'='*50}")